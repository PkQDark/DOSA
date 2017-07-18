from django.shortcuts import render, HttpResponse, HttpResponseRedirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required

from openpyxl import Workbook
from datetime import datetime
from pytz import timezone
import socket

from company_admin.models import KeyOwner, Database, Cistern, UpDosed, FuelType
from company_admin.forms import DateFilter

kiev = timezone('Europe/Kiev')


# БД отгрузок
@login_required
def dosing(request):
    company = request.user.companyuser.company
    if not company.active:
        messages.add_message(request, messages.ERROR, 'Срок лицензии истек')
        return render(request, 'main/blocked.html')
    start_date = kiev.localize(datetime.utcfromtimestamp(0))
    end_date = kiev.localize(datetime.now())
    filter_name = ''
    filter_car = ''
    filter_fuel = ''
    if request.GET.get('filter_name'):
        filter_name = request.GET.get('filter_name')
    if request.GET.get('filter_car'):
        filter_car = request.GET.get('filter_car')
    if request.GET.get('filter_fuel'):
        filter_fuel = request.GET.get('filter_fuel')
    date_filter = DateFilter(request.GET)
    if date_filter.is_valid():
        if date_filter.cleaned_data['start_date']:
            start_date = kiev.localize(datetime.combine(date_filter.cleaned_data['start_date'], datetime.min.time()))
        if date_filter.cleaned_data['end_date']:
            end_date = kiev.localize(datetime.combine(date_filter.cleaned_data['end_date'], datetime.max.time()))
    db = Database.objects.filter(dev__company=company, date_time__gte=start_date, date_time__lte=end_date,
                                 user__name__icontains=filter_name, user__car__icontains=filter_car,
                                 dev__cistern__fuel__name__icontains=filter_fuel, delete=False)
    if request.GET.get('to_xls'):
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="loads_report.xls"'
        wb = Workbook()
        ws = wb.active
        xls_mas = ['Имя', 'Машина', 'Ключ', 'Отгружено', 'Дата', 'Цистерна', 'Тип жидкости']
        ws.append(xls_mas)
        for entry in db:
            xls_mas = [entry.user.name, entry.user.car, entry.user.keys, entry.dosed, entry.date_time]
            if entry.dev.cistern:
                xls_mas.extend([entry.dev.cistern.name, entry.dev.cistern.fuel.name])
            else:
                xls_mas.extend(['', ''])
            ws.append(xls_mas)
        wb.save(response)
        return response
    return render(request, 'users/user_dosing.html',
                  {'db': db, 'date_filter': date_filter, 'company': company.name, 'cur_user': request.user.username})


# БД ключей
@login_required
def keys(request):
    company = request.user.companyuser.company
    if not company.active:
        messages.add_message(request, messages.ERROR, 'Срок лицензии истек')
        return render(request, 'main/blocked.html')
    filter_name = ''
    filter_car = ''
    if request.GET.get('filter_name'):
        filter_name = request.GET.get('filter_name')
    if request.GET.get('filter_car'):
        filter_car = request.GET.get('filter_car')
    udb = KeyOwner.objects.filter(car__icontains=filter_car, name__icontains=filter_name, company=company)
    if request.GET.get('to_xls'):
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="users_report.xls"'
        wb = Workbook()
        ws = wb.active
        xls_mas = ['Имя', 'Машина', 'Ключ', 'Комментарий']
        ws.append(xls_mas)
        for entry in udb:
            xls_mas = [entry.name, entry.car, entry.keys, entry.comment]
            ws.append(xls_mas)
        wb.save(response)
        return response
    return render(request, 'users/user_keys.html',
                  {'udb': udb, 'cur_user': request.user.username, 'company': company.name})


# Статистика по топливу
@login_required
def fuel_info(request, fuel_id):
    company = request.user.companyuser.company
    if not company.active:
        messages.add_message(request, messages.ERROR, 'Срок лицензии истек')
        return render(request, 'main/blocked.html')
    fuel = FuelType.objects.get(id=fuel_id)
    start_date = kiev.localize(datetime.utcfromtimestamp(0))
    end_date = kiev.localize(datetime.now())
    filter_name = ''
    filter_car = ''
    if request.GET.get('filter_name'):
        filter_name = request.GET.get('filter_name')
    if request.GET.get('filter_car'):
        filter_car = request.GET.get('filter_car')
    date_filter = DateFilter(request.GET)
    if date_filter.is_valid():
        if date_filter.cleaned_data['start_date']:
            start_date = kiev.localize(datetime.combine(date_filter.cleaned_data['start_date'], datetime.min.time()))
        if date_filter.cleaned_data['end_date']:
            end_date = kiev.localize(datetime.combine(date_filter.cleaned_data['end_date'], datetime.max.time()))
    db = Database.objects.filter(date_time__gte=start_date, date_time__lte=end_date, dev__company=company,
                                 user__name__icontains=filter_name, user__car__icontains=filter_car,
                                 dev__cistern__fuel=fuel, delete=False)
    if request.GET.get('to_xls'):
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="loads_report.xls"'
        wb = Workbook()
        ws = wb.active
        xls_mas = ['Имя', 'Машина', 'Ключ', 'Отгружено', 'Дата', 'Оставшийся объем', 'Цистерна']
        ws.append(xls_mas)
        for entry in db:
            xls_mas = [entry.user.name, entry.user.car, entry.user.keys, entry.dosed, entry.date_time,
                       entry.fuel_volume, entry.dev.cistern.name]
            ws.append(xls_mas)
        wb.save(response)
        return response
    return render(request, 'users/user_fuel_info.html',
                  {'db': db, 'date_filter': date_filter, 'cur_user': request.user.username, 'company': company.name})


# Перечень резервуаров
@login_required
def cistern_list(request):
    company = request.user.companyuser.company
    if not company.active:
        messages.add_message(request, messages.ERROR, 'Срок лицензии истек')
        return render(request, 'main/blocked.html')
    cists = Cistern.objects.filter(dev__company=company)
    if len(cists) == 0:
        messages.add_message(request, messages.WARNING, 'Нет зарегистрированных резервуаров. '
                                                        'Обратитесь к администратору компании')
        if not FuelType.objects.exists():
            messages.add_message(request, messages.WARNING, 'Не заданы типы топлива. '
                                                            'Обратитесь к администратору компании')
            return HttpResponseRedirect('/user/')
        return HttpResponseRedirect('/user/')
    percents = []
    cur_vols = []
    fuels = FuelType.objects.all()
    if request.GET.get('select_fuel'):
        cists = Cistern.objects.filter(fuel__name=request.GET.get('select_fuel'), dev__company=company)
    for c in cists:
        try:
            db = Database.objects.filter(dev__cistern=c).latest('date_time')
            cur_volume = db.cistern_volume
            cur_vols.append(cur_volume)
        except Database.DoesNotExist:
            cur_volume = c.start_volume
            cur_vols.append(c.start_volume)
        perc = int(cur_volume / c.max_volume * 100 + 5 - (cur_volume / c.max_volume * 100 + 5) % 10)
        percents.append(perc)
    if request.GET.get('refr_log'):
        dev_id = request.GET.get('refr_log')
        sock = socket.socket()
        sock.connect(('195.12.59.37', 9090))
        comand = b'view_log\r'
        n = b'\n'
        space = b' '
        dev_id = dev_id.encode()
        to_send = dev_id + space + comand + n
        sock.send(to_send)
        sock.close()
        # Обновление БД отгрузок
    if request.GET.get('refr_keys'):
        dev_id = request.GET.get('refr_keys')
        sock = socket.socket()
        sock.connect(('195.12.59.37', 9090))
        comand = b'view_keys\r'
        n = b'\n'
        space = b' '
        dev_id = dev_id.encode()
        to_send = dev_id + space + comand + n
        sock.send(to_send)
        sock.close()
        # Обновление БД ключей
    return render(request, 'users/user_cisterns.html',
                  {'cists': zip(cists, cur_vols, percents), 'cur_user': request.user.username, 'fuels': fuels,
                   'company': company.name})


# Статистика резервуара
@login_required
def cistern_info(request, cist_id):
    company = request.user.companyuser.company
    if not company.active:
        messages.add_message(request, messages.ERROR, 'Срок лицензии истек')
        return render(request, 'main/blocked.html')
    nav = "downdosed"
    cist = Cistern.objects.get(id=cist_id)
    start_date = kiev.localize(datetime.utcfromtimestamp(0))
    end_date = kiev.localize(datetime.now())
    filter_name = ''
    filter_car = ''
    if request.GET.get('filter_name'):
        filter_name = request.GET.get('filter_name')
    if request.GET.get('filter_car'):
        filter_car = request.GET.get('filter_car')
    date_filter = DateFilter(request.GET)
    if date_filter.is_valid():
        if date_filter.cleaned_data['start_date']:
            start_date = kiev.localize(datetime.combine(date_filter.cleaned_data['start_date'], datetime.min.time()))
        if date_filter.cleaned_data['end_date']:
            end_date = kiev.localize(datetime.combine(date_filter.cleaned_data['end_date'], datetime.max.time()))
    if request.GET.get('nav'):
        nav = request.GET.get('nav')
    downdosed = Database.objects.filter(date_time__gte=start_date, date_time__lte=end_date, delete=False,
                                        user__name__icontains=filter_name, user__car__icontains=filter_car,
                                        dev__cistern=cist, dev__company=company)
    updosed = UpDosed.objects.filter(date_time__gte=start_date, date_time__lte=end_date, dev__cistern=cist,
                                     dev__company=company)
    if request.GET.get('to_xls') and nav == "downdosed":
        response = HttpResponse(content_type='application/vnd.ms-excel')
        filename = 'downdosed_' + cist_id + '.xls"'
        response['Content-Disposition'] = 'attachment; filename="' + filename
        wb = Workbook()
        ws = wb.active
        xls_mas = ['Имя', 'Машина', 'Ключ', 'Отгружено', 'Дата', 'Объем в цистерне', 'Добавочный объем']
        ws.append(xls_mas)
        for entry in downdosed:
            xls_mas = [entry.user.name, entry.user.car, entry.user.keys,
                       entry.dosed, entry.date_time, entry.cistern_volume, entry.add]
            ws.append(xls_mas)
        wb.save(response)
        return response
    if request.GET.get('to_xls') and nav == "updosed":
        response = HttpResponse(content_type='application/vnd.ms-excel')
        filename = 'updosed_' + cist_id + '.xls"'
        response['Content-Disposition'] = 'attachment; filename="' + filename
        wb = Workbook()
        ws = wb.active
        xls_mas = ['Имя', 'Фамилия', 'Объем', 'Дата', 'Комментарий']
        ws.append(xls_mas)
        for entry in updosed:
            xls_mas = [entry.user.user.first_name, entry.user.user.last_name, entry.volume, entry.date_time, entry.comment]
            ws.append(xls_mas)
        wb.save(response)
        return response
    return render(request, 'users/user_cistern_info.html',
                  {'cur_user': request.user.username, 'nav': nav, 'date_filter': date_filter, 'downdosed': downdosed,
                   'updosed': updosed, 'company': company.name})
