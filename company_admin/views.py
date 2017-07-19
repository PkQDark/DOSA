from django.shortcuts import render, HttpResponse, HttpResponseRedirect
from django.contrib import messages
from django.utils.datastructures import MultiValueDictKeyError
from django.contrib.auth.models import User
from django.contrib.auth.forms import UserCreationForm, SetPasswordForm
from django.contrib.auth.decorators import user_passes_test
from django.forms.formsets import formset_factory
from django.db import IntegrityError
from django.db.models import Sum
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

from datetime import datetime
from pytz import timezone
from openpyxl import Workbook
import time
from .models import KeyOwner, Database, Cistern, UpDosed, FuelType, DayLimit, WeekLimit, MonthLimit
from .forms import DateFilter, LimitsForm, EditKeyOwnerForm, EditDjangoUserForm, FuelForm, BaseFuelFormSet, CisternForm, \
    AddUpDosedForm
from local_admin.models import CompanyUser, Device
from local_admin.forms import AddSystemUserForm, AddIDForm
import socket

kiev = timezone('Europe/Kiev')


def recalc(request, cist, fuels):
    cist_db = Database.objects.filter(dev__cistern=cist).order_by('date_time')
    if len(cist_db):
        cist_db[0].cistern_volume = cist.start_volume - cist_db[0].dosed + cist_db[0].add
        cist_db[0].save()
        cur_cist_volume = cist_db[0].cistern_volume
        for cist_load in cist_db[1:]:
            cist_load.cistern_volume = cur_cist_volume - cist_load.dosed + cist_load.add
            cur_cist_volume = cist_load.cistern_volume
            cist_load.save()
    for f in fuels:
        fuel_db = Database.objects.filter(dev__cistern__fuel=f).order_by('date_time')
        if len(fuel_db):
            fuel_vol = Cistern.objects.filter(fuel=f).aggregate(Sum('start_volume'))['start_volume__sum']
            fuel_db[0].fuel_volume = fuel_vol - fuel_db[0].dosed + fuel_db[0].add
            fuel_db[0].save()
            fuel_vol = fuel_db[0].fuel_volume
            for fuel_load in fuel_db[1:]:
                fuel_load.fuel_volume = fuel_vol - fuel_load.dosed + fuel_load.add
                fuel_vol = fuel_load.fuel_volume
                fuel_load.save()


# БД отгрузок
@user_passes_test(lambda u: u.is_active)
@user_passes_test(lambda u: u.is_staff)
def dosing(request):
    company = request.user.companyuser.company
    if not company.active:
        messages.add_message(request, messages.ERROR, 'Срок лицензии истек')
        return render(request, 'main/blocked.html')
    start_date = kiev.localize(datetime.utcfromtimestamp(0))
    end_date = kiev.localize(datetime.now())
    filter_name = ''
    filter_car = ''
    filter_fuel = ''
    if request.GET.get('filter_name'):
        filter_name = request.GET.get('filter_name')
    if request.GET.get('filter_car'):
        filter_car = request.GET.get('filter_car')
    if request.GET.get('filter_fuel'):
        filter_fuel = request.GET.get('filter_fuel')
    date_filter = DateFilter(request.GET)
    if date_filter.is_valid():
        if date_filter.cleaned_data['start_date']:
            start_date = kiev.localize(datetime.combine(date_filter.cleaned_data['start_date'], datetime.min.time()))
        if date_filter.cleaned_data['end_date']:
            end_date = kiev.localize(datetime.combine(date_filter.cleaned_data['end_date'], datetime.max.time()))
    downdosed = Database.objects.filter(dev__company=company, date_time__gte=start_date, date_time__lte=end_date,
                                        user__name__icontains=filter_name, user__car__icontains=filter_car,
                                        dev__cistern__fuel__name__icontains=filter_fuel, delete=False)
    if request.GET.get('delete'):
        downdosed.update(delete=True)
    if request.GET.get('to_xls'):
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="loads_report.xls"'
        wb = Workbook()
        ws = wb.active
        xls_mas = ['Имя', 'Машина', 'Ключ', 'Отгружено', 'Дата', 'Цистерна', 'Тип жидкости']
        ws.append(xls_mas)
        for entry in downdosed:
            xls_mas = [entry.user.name, entry.user.car, entry.user.keys, entry.dosed, entry.date_time]
            if entry.dev.cistern:
                xls_mas.extend([entry.dev.cistern.name, entry.dev.cistern.fuel.name])
            else:
                xls_mas.extend(['', ''])
            ws.append(xls_mas)
        wb.save(response)
        return response
    # Пагинация
    db_paginator = Paginator(downdosed, 25)
    db_page = request.GET.get('db_page')
    try:
        db = db_paginator.page(db_page)
    except PageNotAnInteger:
        db = db_paginator.page(1)
    except EmptyPage:
        db = db_paginator.page(db_paginator.num_pages)
    return render(request, 'company_admin/admin_dosing.html',
                  {'db': db, 'date_filter': date_filter, 'company': company.name, 'cur_user': request.user.username})


# БД ключей
@user_passes_test(lambda u: u.is_active)
@user_passes_test(lambda u: u.is_staff)
def keys(request):
    company = request.user.companyuser.company
    if not company.active:
        messages.add_message(request, messages.ERROR, 'Срок лицензии истек')
        return render(request, 'main/blocked.html')
    filter_name = ''
    filter_car = ''
    if request.GET.get('filter_name'):
        filter_name = request.GET.get('filter_name')
    if request.GET.get('filter_car'):
        filter_car = request.GET.get('filter_car')
    udb = KeyOwner.objects.filter(car__icontains=filter_car, name__icontains=filter_name, company=company)
    if request.GET.get('to_xls'):
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="users_report.xls"'
        wb = Workbook()
        ws = wb.active
        xls_mas = ['Имя', 'Машина', 'Ключ', 'Комментарий']
        ws.append(xls_mas)
        for entry in udb:
            xls_mas = [entry.name, entry.car, entry.keys, entry.comment]
            ws.append(xls_mas)
        wb.save(response)
        return response
    return render(request, 'company_admin/admin_keys.html',
                  {'udb': udb, 'cur_user': request.user.username, 'company': company.name})


# Редактирование БД ключей
@user_passes_test(lambda u: u.is_active)
@user_passes_test(lambda u: u.is_staff)
def edit_key(request, key_id):
    company = request.user.companyuser.company
    if not company.active:
        messages.add_message(request, messages.ERROR, 'Срок лицензии истек')
        return render(request, 'main/blocked.html')
    owner = KeyOwner.objects.get(id=key_id)
    limits = {}
    if hasattr(owner, 'daylimit'):
        limits['day_limit'] = owner.daylimit.start_volume
    if hasattr(owner, 'weeklimit'):
        limits['week_limit'] = owner.weeklimit.start_volume
    if hasattr(owner, 'monthlimit'):
        limits['month_limit'] = owner.monthlimit.start_volume
    if request.method == 'POST':
        edit_form = EditKeyOwnerForm(request.POST, instance=owner)
        limit_form = LimitsForm(request.POST, initial=limits)
        if edit_form.is_valid():
            edit_form.save()
        if limit_form.is_valid():
            t_date = datetime.now().date()
            today = kiev.localize(datetime.combine(t_date, datetime.min.time()))
            if limit_form.cleaned_data['day_limit']:
                if 'day_limit' not in limits:
                    dl = DayLimit(start_volume=limit_form.cleaned_data['day_limit'],
                                  start_date=today, key_owner=owner)
                    dl.save()
                elif limit_form.cleaned_data['day_limit'] != limits['day_limit']:
                    dl = owner.daylimit
                    dl.start_date = today
                    dl.start_volume = limit_form.cleaned_data['day_limit']
                    dl.save()
            if limit_form.cleaned_data['week_limit']:
                if 'week_limit' not in limits:
                    wl = WeekLimit(start_volume=limit_form.cleaned_data['week_limit'],
                                   start_date=today, key_owner=owner)
                    wl.save()
                elif limit_form.cleaned_data['week_limit'] != limits['week_limit']:
                    wl = owner.weeklimit
                    wl.start_volume = limit_form.cleaned_data['week_limit']
                    wl.save()
            if limit_form.cleaned_data['month_limit']:
                if 'month_limit' not in limits:
                    ml = MonthLimit(start_volume=limit_form.cleaned_data['month_limit'],
                                    start_date=today, key_owner=owner)
                    ml.save()
                elif limit_form.cleaned_data['month_limit'] != limits['month_limit']:
                    ml = owner.weeklimit
                    ml.start_volume = limit_form.cleaned_data['month_limit']
                    ml.save()
            messages.add_message(request, messages.SUCCESS, 'Информация успешно отредактирована')
            return HttpResponseRedirect('/local-admin/keys/')
    else:
        edit_form = EditKeyOwnerForm(instance=owner)
        limit_form = LimitsForm(initial=limits)
    return render(request, 'company_admin/admin_key_edit.html',
                  {'edit_form': edit_form, 'limit_form': limit_form, 'cur_user': request.user.username,
                   'company': company.name})


# Пользователи системы
@user_passes_test(lambda u: u.is_active)
@user_passes_test(lambda u: u.is_staff)
def users(request):
    company = request.user.companyuser.company
    if not company.active:
        messages.add_message(request, messages.ERROR, 'Срок лицензии истек')
        return render(request, 'main/blocked.html')
    filter_name = ''
    if request.GET.get('filter_name'):
        filter_name = request.GET.get('filter_name')
    udb = User.objects.filter(last_name__icontains=filter_name, companyuser__company=company)
    disable = request.POST.get('user_name')
    if disable:
        User.objects.get(username=disable).delete()
    return render(request, 'company_admin/admin_users.html',
                  {'udb': udb, 'cur_user': request.user.username, 'company': company.name})


# Регистрация пользователя
@user_passes_test(lambda u: u.is_active)
@user_passes_test(lambda u: u.is_staff)
def add_local_user(request):
    company = request.user.companyuser.company
    if not company.active:
        messages.add_message(request, messages.ERROR, 'Срок лицензии истек')
        return render(request, 'main/blocked.html')
    if request.method == 'POST':
        general = UserCreationForm(request.POST)
        additional = AddSystemUserForm(request.POST)
        if general.is_valid() and additional.is_valid():
            e_mail = additional.cleaned_data['email']
            try:
                User.objects.get(email=e_mail)
                messages.add_message(request, messages.ERROR,
                                     'Пользователь с почтовым ящиком ' + e_mail + ' уже зарегистрирован')
            except User.DoesNotExist:
                user = general.save()
                user.first_name = additional.cleaned_data['first_name']
                user.last_name = additional.cleaned_data['last_name']
                user.email = e_mail
                user.save()
                CompanyUser.objects.create(user=user, company=company)
                messages.add_message(request, messages.SUCCESS,
                                     'Пользователь успешно зарегистрирован')
                return HttpResponseRedirect('/local-admin/users/')
    else:
        general = UserCreationForm()
        additional = AddSystemUserForm()
    return render(request, 'company_admin/admin_user_add.html',
                  {'general': general, 'additional': additional, 'cur_user': request.user.username, 'company': company.name})


# Редактирование пользователя
@user_passes_test(lambda u: u.is_active)
@user_passes_test(lambda u: u.is_staff)
def edit_local_user(request, user_id):
    company = request.user.companyuser.company
    if not company.active:
        messages.add_message(request, messages.ERROR, 'Срок лицензии истек')
        return render(request, 'main/blocked.html')
    u = User.objects.get(id=user_id)
    if request.method == 'POST':
        edit_form = EditDjangoUserForm(request.POST, instance=u)
        set_passwd = SetPasswordForm(user=u)
        if request.POST.get('ch_user'):
            if edit_form.is_valid():
                edit_form.save()
                messages.add_message(request, messages.SUCCESS, 'Информация о пользователе успешно отредактирована.')
        if request.POST.get('set_passwd'):
            set_passwd = SetPasswordForm(data=request.POST, user=u)
            if set_passwd.is_valid():
                set_passwd.save()
                messages.add_message(request, messages.SUCCESS, 'Пароль успешно изменен.')
    else:
        edit_form = EditDjangoUserForm(instance=u)
        set_passwd = SetPasswordForm(user=u)
    return render(request, 'company_admin/admin_user_edit.html',
                  {'edit_form': edit_form, 'cur_user': request.user.username, 'e_user': u.username,
                   'set_passwd': set_passwd, 'company': company.name})


# Добавление типа топлива
@user_passes_test(lambda u: u.is_active)
@user_passes_test(lambda u: u.is_staff)
def add_fuel(request):
    company = request.user.companyuser.company
    if not company.active:
        messages.add_message(request, messages.ERROR, 'Срок лицензии истек')
        return render(request, 'main/blocked.html')
    FuelFormSet = formset_factory(FuelForm, formset=BaseFuelFormSet, can_delete=False)
    if request.method == 'POST':
        fuel_formset = FuelFormSet(request.POST)
        if fuel_formset.is_valid():
            for fuel_form in fuel_formset:
                new_fuel = FuelType(name=fuel_form.cleaned_data.get('name'), company=company)
                if fuel_form.cleaned_data.get('comment'):
                    new_fuel.comment = fuel_form.cleaned_data.get('comment')
                try:
                    new_fuel.save()
                except IntegrityError:
                    messages.add_message(request, messages.ERROR, 'Данный вид топлива уже зарегистрирован.')
        return HttpResponseRedirect('/local-admin/fuels/')
    else:
        fuel_formset = FuelFormSet()
    return render(request, 'company_admin/admin_fuel_add.html',
                  {'cur_user': request.user.username, 'fuel_formset': fuel_formset, 'company': company.name})


# Типы топлива
@user_passes_test(lambda u: u.is_active)
@user_passes_test(lambda u: u.is_staff)
def fuels_list(request):
    company = request.user.companyuser.company
    if not company.active:
        messages.add_message(request, messages.ERROR, 'Срок лицензии истек')
        return render(request, 'main/blocked.html')
    fuels = FuelType.objects.filter(company=company)
    if request.method == 'POST':
        edit_fuel = FuelForm(request.POST)
        if edit_fuel.is_valid():
            fuel = FuelType.objects.get(id=request.POST.get('fuel_id'))
            fuel.name = edit_fuel.cleaned_data['name']
            if edit_fuel.cleaned_data['comment']:
                fuel.comment = edit_fuel.cleaned_data['comment']
            try:
                fuel.save()
                messages.add_message(request, messages.SUCCESS, 'Информация успешно отредактирована.')
            except IntegrityError:
                messages.add_message(request, messages.ERROR, 'Данный вид топлива уже зарегистрирован.')
    else:
        edit_fuel = FuelForm()
    if request.GET.get('del'):
        fuel_del = FuelType.objects.get(id=request.GET.get('del'))
        if len(fuel_del.cistern_set.all()):
            messages.add_message(request, messages.ERROR, 'Удаление невозможно. '
                                                          'Данный тип топлива задействован в резервуаре.')
        else:
            fuel_del.delete()
    return render(request, 'company_admin/admin_fuels.html',
                  {'cur_user': request.user.username, 'fuels': fuels, 'edit_fuel': edit_fuel, 'company': company.name})


# Статистика по топливу
@user_passes_test(lambda u: u.is_active)
@user_passes_test(lambda u: u.is_staff)
def fuel_info(request, fuel_id):
    company = request.user.companyuser.company
    if not company.active:
        messages.add_message(request, messages.ERROR, 'Срок лицензии истек')
        return render(request, 'main/blocked.html')
    fuel = FuelType.objects.get(id=fuel_id)
    start_date = kiev.localize(datetime.utcfromtimestamp(0))
    end_date = kiev.localize(datetime.now())
    filter_name = ''
    filter_car = ''
    if request.GET.get('filter_name'):
        filter_name = request.GET.get('filter_name')
    if request.GET.get('filter_car'):
        filter_car = request.GET.get('filter_car')
    date_filter = DateFilter(request.GET)
    if date_filter.is_valid():
        if date_filter.cleaned_data['start_date']:
            start_date = kiev.localize(datetime.combine(date_filter.cleaned_data['start_date'], datetime.min.time()))
        if date_filter.cleaned_data['end_date']:
            end_date = kiev.localize(datetime.combine(date_filter.cleaned_data['end_date'], datetime.max.time()))
    db = Database.objects.filter(date_time__gte=start_date, date_time__lte=end_date, dev__company=company,
                                 user__name__icontains=filter_name, user__car__icontains=filter_car,
                                 dev__cistern__fuel=fuel, delete=False)
    if request.GET.get('delete'):
        db.update(delete=True)
    if request.GET.get('to_xls'):
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="loads_report.xls"'
        wb = Workbook()
        ws = wb.active
        xls_mas = ['Имя', 'Машина', 'Ключ', 'Отгружено', 'Дата', 'Оставшийся объем', 'Цистерна']
        ws.append(xls_mas)
        for entry in db:
            xls_mas = [entry.user.name, entry.user.car, entry.user.keys, entry.dosed, entry.date_time,
                       entry.fuel_volume, entry.dev.cistern.name]
            ws.append(xls_mas)
        wb.save(response)
        return response
    # Пагинация
    db_paginator = Paginator(db, 25)
    db_page = request.GET.get('db_page')
    try:
        downdosed = db_paginator.page(db_page)
    except PageNotAnInteger:
        downdosed = db_paginator.page(1)
    except EmptyPage:
        downdosed = db_paginator.page(db_paginator.num_pages)
    return render(request, 'company_admin/admin_fuel_info.html',
                  {'db': downdosed, 'date_filter': date_filter, 'cur_user': request.user.username,
                   'company': company.name})


# Добавление резервуара
@user_passes_test(lambda u: u.is_active)
@user_passes_test(lambda u: u.is_staff)
def cistern_add(request):
    company = request.user.companyuser.company
    if not company.active:
        messages.add_message(request, messages.ERROR, 'Срок лицензии истек')
        return render(request, 'main/blocked.html')
    fuel_types = FuelType.objects.filter(company=company)
    if len(fuel_types) == 0:
        messages.add_message(request, messages.WARNING, 'Не заданы типы топлива, добавление резервуара невозможно')
        return HttpResponseRedirect('/local-admin/fuels/add-fuel/')
    free_devs = Device.objects.filter(owned=False, company=company)
    if len(free_devs) == 0:
        messages.add_message(request, messages.WARNING, 'Нет незадействованных устройств, добавление резервуара невозможно')
        return HttpResponseRedirect('/local-admin/cisterns/')
    if request.method == 'POST':
        add_cist = CisternForm(request.POST)
        if add_cist.is_valid():
            try:
                fuel = FuelType.objects.get(name=request.POST['select_fuel'])
                dev = Device.objects.get(dev_id=request.POST['select_dev'])
                dev.owned = True
                dev.save()
                cist = add_cist.save(commit=False)
                cist.fuel = fuel
                cist.dev = dev
                cist.save()
                messages.add_message(request, messages.SUCCESS, 'Резервуар успешно добавлен')
                return HttpResponseRedirect('/local-admin/cisterns/')
            except ValueError:
                messages.add_message(request, messages.ERROR, 'Начальное значение не может превышать максимальное')
                return HttpResponseRedirect('/local-admin/cisterns/add-cistern')
    else:
        add_cist = CisternForm()
    return render(request, 'company_admin/admin_cistern_add.html',
                  {"add_cist": add_cist, 'devs': free_devs, 'fuel_types': fuel_types,
                   'cur_user': request.user.username, 'company': company.name})


def send_request_to_data(dev_id, comand, request):
    sock = socket.socket()
    port = Device.objects.get(dev_id=dev_id)
    dev_port = port.port
    sock.connect(('195.12.59.37', dev_port))
    n = b'\n'
    space = b' '
    dev_id = dev_id.encode()
    to_send = dev_id + space + comand + n
    sock.send(to_send)
    input_text = b''
    time.sleep(1)
    while not input_text.endswith(b'\r'):
        input_text += sock.recv(1024)
    if input_text == b'yes\r':
        messages.add_message(request, messages.INFO, 'Устройство найдено, информация вскоре будет обновлена')
    elif input_text == b'no\r':
        messages.add_message(request, messages.INFO,
                             'Устройство не найдено, попробуйте позже, или проверьте соединение')
    else:
        messages.add_message(request, messages.INFO, input_text.decode())
    sock.close()


# Перечень резервуаров
@user_passes_test(lambda u: u.is_active)
@user_passes_test(lambda u: u.is_staff)
def cistern_list(request):
    company = request.user.companyuser.company
    if not company.active:
        messages.add_message(request, messages.ERROR, 'Срок лицензии истек')
        return render(request, 'main/blocked.html')
    cists = Cistern.objects.filter(dev__company=company)
    if len(cists) == 0:
        messages.add_message(request, messages.WARNING,
                             'Нет зарегистрированных резервуаров. Добавьте резервуар для продолжения работы.')
        if not FuelType.objects.exists():
            messages.add_message(request, messages.WARNING, 'Задайте типы топлива для продолжения работы.')
            return HttpResponseRedirect('/local-admin/fuels/add-fuel/')
        return HttpResponseRedirect('/local-admin/cisterns/add-cistern/')
    percents = []
    cur_vols = []
    fuels = FuelType.objects.all()
    if request.GET.get('select_fuel'):
        cists = Cistern.objects.filter(fuel__name=request.GET.get('select_fuel'), dev__company=company)
    for c in cists:
        try:
            db = Database.objects.filter(dev__cistern=c).latest('date_time')
            cur_volume = db.cistern_volume
            cur_vols.append(cur_volume)
        except Database.DoesNotExist:
            cur_volume = c.start_volume
            cur_vols.append(c.start_volume)
        perc = int(cur_volume / c.max_volume * 100 + 5 - (cur_volume / c.max_volume * 100 + 5) % 10)
        percents.append(perc)
    if request.GET.get('refr_log'):
        dev_id = request.GET.get('refr_log')
        comand = b'view_log\r'
        send_request_to_data(dev_id, comand, request)
        # Обновление бд отгрузок
    if request.GET.get('refr_keys'):
        dev_id = request.GET.get('refr_keys')
        comand = b'view_keys\r'
        send_request_to_data(dev_id, comand, request)
        # Обновление БД ключей
    return render(request, 'company_admin/admin_cisterns.html',
                  {'cists': zip(cists, cur_vols, percents), 'cur_user': request.user.username, 'fuels': fuels,
                   'company': company.name})


# Редактирование резервуара
@user_passes_test(lambda u: u.is_active)
@user_passes_test(lambda u: u.is_staff)
def cistern_edit(request, cist_id):
    company = request.user.companyuser.company
    if not company.active:
        messages.add_message(request, messages.ERROR, 'Срок лицензии истек')
        return render(request, 'main/blocked.html')
    c = Cistern.objects.get(id=cist_id)
    free_devs = Device.objects.filter(owned=False)
    fuel_types = FuelType.objects.exclude(id=c.fuel.id)
    if len(fuel_types) == 0:
        fuel_types = None
    if len(free_devs) == 0:
        free_devs = None
    if request.method == 'POST':
        edit_cist = CisternForm(request.POST, instance=c)
        try:
            if edit_cist.is_valid():
                cist = edit_cist.save(commit=False)
                fuel_recalc = [c.fuel]
                if request.POST['select_fuel'] != c.fuel.name:
                    cist.fuel = FuelType.objects.get(name=request.POST['select_fuel'], company=company)
                    fuel_recalc.append(c.fuel)
                if request.POST['select_dev'] != c.dev.dev_id:
                    dev = c.dev
                    dev.owned = False
                    dev.save()
                    new_dev = Device.objects.get(dev_id=request.POST['select_dev'], company=company)
                    new_dev.owned = True
                    cist.dev = new_dev
                cist.save()
                if Database.objects.filter(dev__cistern__fuel__in=fuel_recalc).count() > 0:
                    recalc(request, c, fuel_recalc)
                messages.add_message(request, messages.SUCCESS, 'Значения успешно изменены')
                return HttpResponseRedirect('/local-admin/cisterns/')
        except ValueError:
            messages.add_message(request, messages.ERROR, 'Начальное значение не может превышать максимальное')
            return HttpResponseRedirect('/local-admin/cisterns/')
    else:
        edit_cist = CisternForm(instance=c)
    return render(request, 'company_admin/admin_cistern_edit.html',
                  {"edit_cist": edit_cist, 'cur_dev': c.dev.dev_id, 'free_devs': free_devs, 'cur_fuel': c.fuel.name,
                   'fuel_types': fuel_types, 'cur_user': request.user.username, 'company': company.name})


# Статистика резервуара
@user_passes_test(lambda u: u.is_active)
@user_passes_test(lambda u: u.is_staff)
def cistern_info(request, cist_id):
    company = request.user.companyuser.company
    if not company.active:
        messages.add_message(request, messages.ERROR, 'Срок лицензии истек')
        return render(request, 'main/blocked.html')
    nav = "downdosed"
    cist = Cistern.objects.get(id=cist_id)
    start_date = kiev.localize(datetime.utcfromtimestamp(0))
    end_date = kiev.localize(datetime.now())
    filter_name = ''
    filter_car = ''
    if request.GET.get('filter_name'):
        filter_name = request.GET.get('filter_name')
    if request.GET.get('filter_car'):
        filter_car = request.GET.get('filter_car')
    date_filter = DateFilter(request.GET)
    if date_filter.is_valid():
        if date_filter.cleaned_data['start_date']:
            start_date = kiev.localize(datetime.combine(date_filter.cleaned_data['start_date'], datetime.min.time()))
        if date_filter.cleaned_data['end_date']:
            end_date = kiev.localize(datetime.combine(date_filter.cleaned_data['end_date'], datetime.max.time()))
    if request.GET.get('nav'):
        nav = request.GET.get('nav')
    downdosed = Database.objects.filter(date_time__gte=start_date, date_time__lte=end_date, delete=False,
                                        user__name__icontains=filter_name, user__car__icontains=filter_car,
                                        dev__cistern=cist, dev__company=company)
    updosed = UpDosed.objects.filter(date_time__gte=start_date, date_time__lte=end_date, dev__cistern=cist,
                                     dev__company=company)
    recovery = Database.objects.filter(date_time__gte=start_date, date_time__lte=end_date, delete=True,
                                       user__name__icontains=filter_name, user__car__icontains=filter_car,
                                       dev__cistern=cist, dev__company=company)
    if request.method == 'POST':
        add_updosed = AddUpDosedForm(request.POST)
        if add_updosed.is_valid():
            if add_updosed.cleaned_data['volume']:
                load = Database.objects.get(id=request.POST.get('load_id'))
                if add_updosed.cleaned_data['volume'] + load.cistern_volume > load.cistern.max_volume:
                    messages.add_message(request, messages.WARNING, 'Текущий объем не может быть больше максимального')
                    return HttpResponseRedirect('/local-admin/cisterns/')
                add = UpDosed(user=request.user.companyuser, dev=cist.dev, date_time=load.date_time,
                              volume=add_updosed.cleaned_data['volume'])
                if add_updosed.cleaned_data['comment']:
                    add.comment = add_updosed.cleaned_data['comment']
                add.save()
                load.add += add_updosed.cleaned_data['volume']
                load.cistern_volume += add_updosed.cleaned_data['volume']
                load.fuel_volume += add_updosed.cleaned_data['volume']
                load.save()
                next_cist_dosings = Database.objects.filter(dev__cistern=cist, dev__company=company,
                                                            date_time__gt=load.date_time).order_by('date_time')
                if len(next_cist_dosings):
                    previous_cist_volume = load.cistern_volume
                    for db in next_cist_dosings:
                        db.cistern_volume = previous_cist_volume - db.dosed + db.add
                        previous_cist_volume = db.cistern_volume
                        db.save()
                next_dosings = Database.objects.filter(date_time__gt=load.date_time, dev__company=company,
                                                       dev__cistern__fuel=load.user.cistern.fuel).order_by('date_time')
                if len(next_dosings):
                    previous_fuel_volume = load.fuel_volume
                    for db in next_dosings:
                        db.fuel_volume = previous_fuel_volume - db.dosed + db.add
                        previous_fuel_volume = db.fuel_volume
                        db.save()
                messages.add_message(request, messages.SUCCESS, 'Загрузка успешно добавлена')
        add_keys_form = AddIDForm(request.POST, request.FILES)
        if len(request.FILES) != 0:
            try:
                key_doc = request.FILES['key_file'].read().decode().split()
            except MultiValueDictKeyError:
                messages.add_message(request, messages.ERROR, 'Ошибка файла')
                return HttpResponseRedirect('/local-admin/keys/')
            messages.add_message(request, messages.INFO, 'Выполняется загрузка ключей. Это может занять некоторое время')
            dev_id = cist.dev.dev_id
            # Запись ключей на устройство
            sock = socket.socket()
            dev_port = Device.objects.get(dev_id).port
            sock.connect(('195.12.59.37', dev_port))
            comand = b'import_keys\r'
            n = b'\n'
            space = b' '
            dev_id = dev_id.encode()
            to_send = dev_id + space + comand
            for i in key_doc:
                to_send += i.encode() + space
            to_send += n
            sock.send(to_send)
            input_text = b''
            while not input_text.endswith(b'\r'):
                input_text += sock.recv(1024)
            if input_text == b'yes\r':
                messages.add_message(request, messages.INFO, 'Устройство найдено, ключи скоро будут добавлены на устройство')
            elif input_text == b'no\r':
                messages.add_message(request, messages.INFO, 'Устройство не найдено, повторите попытку позже, или проверьте соединение с устройством')
            else:
                messages.add_message(request, messages.INFO, input_text.decode())
            sock.close()

    else:
        add_keys_form = AddIDForm()
        add_updosed = AddUpDosedForm()
    if request.GET.get('delete'):
        downdosed.update(delete=True)
    if request.GET.get('recover'):
        recovery.update(delete=False)
    if request.GET.get('to_xls') and nav == "downdosed":
        response = HttpResponse(content_type='application/vnd.ms-excel')
        filename = 'downdosed_' + cist_id + '.xls"'
        response['Content-Disposition'] = 'attachment; filename="' + filename
        wb = Workbook()
        ws = wb.active
        xls_mas = ['Имя', 'Машина', 'Ключ', 'Отгружено', 'Дата', 'Объем в цистерне', 'Добавочный объем']
        ws.append(xls_mas)
        for entry in downdosed:
            xls_mas = [entry.user.name, entry.user.car, entry.user.keys,
                       entry.dosed, entry.date_time, entry.cistern_volume, entry.add]
            ws.append(xls_mas)
        wb.save(response)
        return response
    if request.GET.get('to_xls') and nav == "updosed":
        response = HttpResponse(content_type='application/vnd.ms-excel')
        filename = 'updosed_' + cist_id + '.xls"'
        response['Content-Disposition'] = 'attachment; filename="' + filename
        wb = Workbook()
        ws = wb.active
        xls_mas = ['Имя', 'Фамилия', 'Объем', 'Дата', 'Комментарий']
        ws.append(xls_mas)
        for entry in updosed:
            xls_mas = [entry.user.user.first_name, entry.user.user.last_name, entry.volume, entry.date_time, entry.comment]
            ws.append(xls_mas)
        wb.save(response)
        return response
    # Пагинация
    db_paginator = Paginator(downdosed, 25)
    db_page = request.GET.get('db_page')
    try:
        db = db_paginator.page(db_page)
    except PageNotAnInteger:
        db = db_paginator.page(1)
    except EmptyPage:
        db = db_paginator.page(db_paginator.num_pages)

    ud_paginator = Paginator(updosed, 25)
    ud_page = request.GET.get('ud_page')
    try:
        ud = ud_paginator.page(ud_page)
    except PageNotAnInteger:
        ud = ud_paginator.page(1)
    except EmptyPage:
        ud = ud_paginator.page(ud_paginator.num_pages)

    rec_paginator = Paginator(recovery, 25)
    rec_page = request.GET.get('rec_page')
    try:
        rec = rec_paginator.page(rec_page)
    except PageNotAnInteger:
        rec = rec_paginator.page(1)
    except EmptyPage:
        rec = rec_paginator.page(rec_paginator.num_pages)
    return render(request, 'company_admin/admin_cistern_info.html',
                  {'cur_user': request.user.username, 'nav': nav, 'date_filter': date_filter,
                   'add_keys': add_keys_form, 'downdosed': db, 'updosed': ud, 'recovery': rec,
                   'add_updosed': add_updosed, 'company': company.name})
